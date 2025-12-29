from fastapi import APIRouter, Depends, Query, Response
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from datetime import date, datetime, timedelta
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter
from pydantic import BaseModel
from typing import List, Dict, Any, Optional

from app.database import get_db
from app.models import Admin, Request, RequestItem, User, Exhibitor, Slot, Schedule, Hermano
from app.routers.admin import get_current_admin

router = APIRouter()

# Response models for dashboard
class DashboardStats(BaseModel):
    total_requests: int
    approved_requests: int
    cancelled_requests: int
    total_exhibitors: int
    active_exhibitors: int
    total_slots: int
    occupied_slots: int
    total_users: int
    active_users: int
    occupancy_rate: float
    most_requested_exhibitor: Optional[Dict[str, Any]] = None
    least_requested_exhibitor: Optional[Dict[str, Any]] = None
    peak_day: Optional[Dict[str, Any]] = None
    requests_by_status: Dict[str, int]
    requests_by_exhibitor: List[Dict[str, Any]]
    daily_requests: List[Dict[str, Any]]

@router.get("/dashboard", response_model=DashboardStats)
async def get_dashboard_stats(
    start_date: date = Query(None),
    end_date: date = Query(None),
    admin: Admin = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db)
):
    """Get dashboard statistics for the admin panel"""
    # Si no se especifican fechas, usar el mes actual
    if not start_date:
        start_date = date.today().replace(day=1)
    if not end_date:
        end_date = date.today()
    
    # Total de solicitudes en el rango
    total_requests_result = await db.execute(
        select(func.count(Request.id)).where(
            Request.created_at >= datetime.combine(start_date, datetime.min.time()),
            Request.created_at <= datetime.combine(end_date, datetime.max.time())
        )
    )
    total_requests = total_requests_result.scalar() or 0
    
    # Solicitudes aprobadas
    approved_result = await db.execute(
        select(func.count(Request.id)).where(
            Request.created_at >= datetime.combine(start_date, datetime.min.time()),
            Request.created_at <= datetime.combine(end_date, datetime.max.time()),
            RequestItem.status == "approved"
        ).select_from(Request).join(RequestItem)
    )
    approved_requests = len((await db.execute(
        select(RequestItem.id).join(Request).where(
            Request.created_at >= datetime.combine(start_date, datetime.min.time()),
            Request.created_at <= datetime.combine(end_date, datetime.max.time()),
            RequestItem.status == "approved"
        ).distinct()
    )).scalars().all())
    
    # Solicitudes liberadas (status="cancelled" en BD pero se muestran como liberadas)
    cancelled_requests = len((await db.execute(
        select(RequestItem.id).join(Request).where(
            Request.created_at >= datetime.combine(start_date, datetime.min.time()),
            Request.created_at <= datetime.combine(end_date, datetime.max.time()),
            RequestItem.status == "cancelled"
        ).distinct()
    )).scalars().all())
    
    # Total de exhibidores
    total_exhibitors_result = await db.execute(select(func.count(Exhibitor.id)))
    total_exhibitors = total_exhibitors_result.scalar() or 0
    
    # Exhibidores activos
    active_exhibitors_result = await db.execute(
        select(func.count(Exhibitor.id)).where(Exhibitor.is_active == True)
    )
    active_exhibitors = active_exhibitors_result.scalar() or 0
    
    # Total de slots en el rango
    total_slots_result = await db.execute(
        select(func.count(Slot.id)).where(
            Slot.slot_date >= start_date,
            Slot.slot_date <= end_date
        )
    )
    total_slots = total_slots_result.scalar() or 0
    
    # Slots ocupados (con solicitudes aprobadas)
    occupied_slots_result = await db.execute(
        select(func.count(func.distinct(Slot.id))).select_from(Slot).join(RequestItem).where(
            Slot.slot_date >= start_date,
            Slot.slot_date <= end_date,
            RequestItem.status == "approved"
        )
    )
    occupied_slots = occupied_slots_result.scalar() or 0
    
    # Total de usuarios
    total_users_result = await db.execute(select(func.count(User.id)))
    total_users = total_users_result.scalar() or 0
    
    # Usuarios activos (que han hecho solicitudes)
    active_users_result = await db.execute(
        select(func.count(func.distinct(Request.user_id))).where(
            Request.created_at >= datetime.combine(start_date, datetime.min.time()),
            Request.created_at <= datetime.combine(end_date, datetime.max.time())
        )
    )
    active_users = active_users_result.scalar() or 0
    
    # Tasa de ocupación
    occupancy_rate = (occupied_slots / total_slots * 100) if total_slots > 0 else 0
    
    # Exhibidor más solicitado
    most_requested_query = await db.execute(
        select(Exhibitor.id, Exhibitor.name, func.count(RequestItem.id).label('count'))
        .join(Slot, Slot.exhibitor_id == Exhibitor.id)
        .join(RequestItem, RequestItem.slot_id == Slot.id)
        .join(Request, Request.id == RequestItem.request_id)
        .where(
            Request.created_at >= datetime.combine(start_date, datetime.min.time()),
            Request.created_at <= datetime.combine(end_date, datetime.max.time()),
            RequestItem.status == "approved"
        )
        .group_by(Exhibitor.id, Exhibitor.name)
        .order_by(func.count(RequestItem.id).desc())
        .limit(1)
    )
    most_requested_row = most_requested_query.first()
    most_requested_exhibitor = {
        "id": str(most_requested_row[0]),
        "name": most_requested_row[1],
        "count": most_requested_row[2]
    } if most_requested_row else None
    
    # Exhibidor menos solicitado
    least_requested_query = await db.execute(
        select(Exhibitor.id, Exhibitor.name, func.count(RequestItem.id).label('count'))
        .join(Slot, Slot.exhibitor_id == Exhibitor.id)
        .outerjoin(RequestItem, RequestItem.slot_id == Slot.id)
        .where(Exhibitor.is_active == True)
        .group_by(Exhibitor.id, Exhibitor.name)
        .order_by(func.count(RequestItem.id).asc())
        .limit(1)
    )
    least_requested_row = least_requested_query.first()
    least_requested_exhibitor = {
        "id": str(least_requested_row[0]),
        "name": least_requested_row[1],
        "count": least_requested_row[2]
    } if least_requested_row else None
    
    # Día pico (día con más solicitudes)
    peak_day_query = await db.execute(
        select(
            func.date(Request.created_at).label('day'),
            func.count(Request.id).label('count')
        )
        .where(
            Request.created_at >= datetime.combine(start_date, datetime.min.time()),
            Request.created_at <= datetime.combine(end_date, datetime.max.time())
        )
        .group_by(func.date(Request.created_at))
        .order_by(func.count(Request.id).desc())
        .limit(1)
    )
    peak_day_row = peak_day_query.first()
    peak_day = {
        "date": str(peak_day_row[0]),
        "count": peak_day_row[1]
    } if peak_day_row else None
    
    # Solicitudes por estado
    status_query = await db.execute(
        select(RequestItem.status, func.count(RequestItem.id))
        .join(Request)
        .where(
            Request.created_at >= datetime.combine(start_date, datetime.min.time()),
            Request.created_at <= datetime.combine(end_date, datetime.max.time())
        )
        .group_by(RequestItem.status)
    )
    requests_by_status = {row[0]: row[1] for row in status_query.all()}
    
    # Top 10 exhibidores por solicitudes
    top_exhibitors_query = await db.execute(
        select(
            Exhibitor.id,
            Exhibitor.name,
            Exhibitor.code,
            func.count(RequestItem.id).label('count')
        )
        .join(Slot, Slot.exhibitor_id == Exhibitor.id)
        .join(RequestItem, RequestItem.slot_id == Slot.id)
        .join(Request, Request.id == RequestItem.request_id)
        .where(
            Request.created_at >= datetime.combine(start_date, datetime.min.time()),
            Request.created_at <= datetime.combine(end_date, datetime.max.time()),
            RequestItem.status == "approved"
        )
        .group_by(Exhibitor.id, Exhibitor.name, Exhibitor.code)
        .order_by(func.count(RequestItem.id).desc())
        .limit(10)
    )
    requests_by_exhibitor = [
        {
            "id": str(row[0]),
            "name": row[1],
            "code": row[2],
            "count": row[3]
        }
        for row in top_exhibitors_query.all()
    ]
    
    # Solicitudes por día (últimos 30 días o rango especificado)
    daily_requests_query = await db.execute(
        select(
            func.date(Request.created_at).label('day'),
            func.count(Request.id).label('count')
        )
        .where(
            Request.created_at >= datetime.combine(start_date, datetime.min.time()),
            Request.created_at <= datetime.combine(end_date, datetime.max.time())
        )
        .group_by(func.date(Request.created_at))
        .order_by(func.date(Request.created_at))
    )
    daily_requests = [
        {
            "date": str(row[0]),
            "count": row[1]
        }
        for row in daily_requests_query.all()
    ]
    
    return DashboardStats(
        total_requests=total_requests,
        approved_requests=approved_requests,
        cancelled_requests=cancelled_requests,
        total_exhibitors=total_exhibitors,
        active_exhibitors=active_exhibitors,
        total_slots=total_slots,
        occupied_slots=occupied_slots,
        total_users=total_users,
        active_users=active_users,
        occupancy_rate=round(occupancy_rate, 2),
        most_requested_exhibitor=most_requested_exhibitor,
        least_requested_exhibitor=least_requested_exhibitor,
        peak_day=peak_day,
        requests_by_status=requests_by_status,
        requests_by_exhibitor=requests_by_exhibitor,
        daily_requests=daily_requests
    )

def create_excel_report(data, headers, title, summary_data=None):
    """Create Excel file from data with enhanced styling"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel limits sheet names to 31 chars
    
    # Styles
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    summary_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    summary_font = Font(bold=True, size=11)
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    current_row = 1
    
    # Agregar resumen si existe
    if summary_data:
        ws.merge_cells(f'A{current_row}:B{current_row}')
        summary_title_cell = ws.cell(row=current_row, column=1, value="📊 RESUMEN")
        summary_title_cell.font = Font(bold=True, size=14, color="0066CC")
        summary_title_cell.alignment = Alignment(horizontal="center")
        current_row += 1
        
        for key, value in summary_data.items():
            ws.cell(row=current_row, column=1, value=key)
            ws.cell(row=current_row, column=2, value=value)
            for col in [1, 2]:
                cell = ws.cell(row=current_row, column=col)
                cell.fill = summary_fill
                cell.font = summary_font if col == 1 else Font(size=11)
                cell.border = border
            current_row += 1
        
        current_row += 2  # Espacio
    
    # Write title
    ws.merge_cells(f'A{current_row}:{chr(64+len(headers))}{current_row}')
    title_cell = ws.cell(row=current_row, column=1, value=title.upper())
    title_cell.font = Font(bold=True, size=14, color="0066CC")
    title_cell.alignment = Alignment(horizontal="center")
    current_row += 1
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    header_row = current_row
    current_row += 1
    
    # Write data
    for row_data in data:
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="center")
        current_row += 1
    
    # Adjust column widths
    for col_idx, column in enumerate(ws.columns, 1):
        max_length = 0
        # Usar get_column_letter para evitar problemas con MergedCell
        column_letter = get_column_letter(col_idx)
        for cell in column:
            try:
                # Saltar celdas fusionadas (MergedCell)
                if hasattr(cell, 'value') and cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max(max_length + 3, 12), 60)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze header row
    ws.freeze_panes = ws[f'A{header_row + 1}']
    
    # Add autofilter
    ws.auto_filter.ref = f'A{header_row}:{chr(64+len(headers))}{current_row-1}'
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

@router.get("/requests")
async def get_requests_report(
    start_date: date = Query(...),
    end_date: date = Query(...),
    admin: Admin = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db)
):
    """Export detailed requests report"""
    # Get all requests with items
    result = await db.execute(
        select(Request, User, RequestItem, Slot, Exhibitor, Schedule)
        .join(User, Request.user_id == User.id)
        .join(RequestItem, RequestItem.request_id == Request.id)
        .join(Slot, RequestItem.slot_id == Slot.id)
        .join(Exhibitor, Slot.exhibitor_id == Exhibitor.id)
        .join(Schedule, Slot.schedule_id == Schedule.id)
        .where(
            Request.created_at >= datetime.combine(start_date, datetime.min.time()),
            Request.created_at <= datetime.combine(end_date, datetime.max.time())
        )
        .order_by(Request.created_at.desc())
    )
    records = result.all()
    
    # Calculate summary
    total_requests = len(set(r[0].id for r in records))
    approved_count = sum(1 for r in records if r[2].status == "approved")
    cancelled_count = sum(1 for r in records if r[2].status == "cancelled")
    
    summary_data = {
        "Periodo": f"{start_date} a {end_date}",
        "Total de Solicitudes": total_requests,
        "Turnos Aprobados": approved_count,
        "Turnos Liberados": cancelled_count,
        "Fecha de Generación": datetime.now().strftime("%Y-%m-%d %H:%M")
    }
    
    data = []
    for req, user, item, slot, exhibitor, schedule in records:
        data.append([
            str(req.id)[:8],
            req.created_at.strftime("%Y-%m-%d %H:%M"),
            user.full_name,
            user.phone,
            exhibitor.code or "N/A",
            exhibitor.name,
            slot.slot_date.strftime("%Y-%m-%d"),
            schedule.start_time.strftime("%H:%M"),
            schedule.end_time.strftime("%H:%M"),
            "LIBERADO" if item.status == "cancelled" else item.status.upper(),
            req.notes or ""
        ])
    
    headers = [
        "ID Solicitud",
        "Fecha Creación",
        "Usuario",
        "Teléfono",
        "Código Exhibidor",
        "Exhibidor",
        "Fecha Turno",
        "Hora Inicio",
        "Hora Fin",
        "Estado",
        "Notas"
    ]
    excel_file = create_excel_report(data, headers, "Reporte Detallado de Solicitudes", summary_data)
    
    return Response(
        content=excel_file.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=solicitudes_detallado_{start_date}_{end_date}.xlsx"}
    )

@router.get("/points")
async def get_exhibitors_report(
    start_date: date = Query(...),
    end_date: date = Query(...),
    admin: Admin = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db)
):
    """Export exhibitors report with detailed statistics"""
    result = await db.execute(select(Exhibitor).order_by(Exhibitor.code, Exhibitor.order))
    exhibitors = result.scalars().all()
    
    data = []
    total_requests_all = 0
    
    for exhibitor in exhibitors:
        # Count approved requests
        approved_result = await db.execute(
            select(func.count(RequestItem.id))
            .join(Slot).join(Request)
            .where(
                Slot.exhibitor_id == exhibitor.id,
                Request.created_at >= datetime.combine(start_date, datetime.min.time()),
                Request.created_at <= datetime.combine(end_date, datetime.max.time()),
                RequestItem.status == "approved"
            )
        )
        approved_count = approved_result.scalar() or 0
        
        # Count liberated requests (status="cancelled" en BD)
        cancelled_result = await db.execute(
            select(func.count(RequestItem.id))
            .join(Slot).join(Request)
            .where(
                Slot.exhibitor_id == exhibitor.id,
                Request.created_at >= datetime.combine(start_date, datetime.min.time()),
                Request.created_at <= datetime.combine(end_date, datetime.max.time()),
                RequestItem.status == "cancelled"
            )
        )
        cancelled_count = cancelled_result.scalar() or 0
        
        # Count total slots created
        total_slots_result = await db.execute(
            select(func.count(Slot.id)).where(
                Slot.exhibitor_id == exhibitor.id,
                Slot.slot_date >= start_date,
                Slot.slot_date <= end_date
            )
        )
        total_slots = total_slots_result.scalar() or 0
        
        # Calculate occupancy rate
        occupancy_rate = (approved_count / total_slots * 100) if total_slots > 0 else 0
        
        total_requests_all += (approved_count + cancelled_count)
        
        data.append([
            exhibitor.code or "N/A",
            exhibitor.name,
            "Activo" if exhibitor.is_active else "Inactivo",
            exhibitor.max_persons_per_slot or 0,
            exhibitor.min_persons_per_slot or 0,
            total_slots,
            approved_count,
            cancelled_count,
            f"{occupancy_rate:.1f}%",
            exhibitor.description or ""
        ])
    
    summary_data = {
        "Periodo": f"{start_date} a {end_date}",
        "Total Exhibidores": len(exhibitors),
        "Exhibidores Activos": sum(1 for e in exhibitors if e.is_active),
        "Total Turnos Solicitados": total_requests_all,
        "Fecha de Generación": datetime.now().strftime("%Y-%m-%d %H:%M")
    }
    
    headers = [
        "Código",
        "Nombre",
        "Estado",
        "Capacidad Máx",
        "Mínimo Req",
        "Slots Totales",
        "Turnos Aprobados",
        "Turnos Liberados",
        "% Ocupación",
        "Descripción"
    ]
    excel_file = create_excel_report(data, headers, "Reporte de Exhibidores", summary_data)
    
    return Response(
        content=excel_file.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=exhibidores_{start_date}_{end_date}.xlsx"}
    )

@router.get("/users")
async def get_users_report(
    start_date: date = Query(...),
    end_date: date = Query(...),
    admin: Admin = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db)
):
    """Export users report with activity statistics"""
    # Get all users who made requests in the date range
    users_result = await db.execute(
        select(User, func.count(Request.id.distinct()).label('total_requests'))
        .join(Request, Request.user_id == User.id)
        .where(
            Request.created_at >= datetime.combine(start_date, datetime.min.time()),
            Request.created_at <= datetime.combine(end_date, datetime.max.time())
        )
        .group_by(User.id, User.full_name, User.phone, User.is_active)
        .order_by(User.full_name)
    )
    users = users_result.all()
    
    data = []
    total_users = len(users)
    total_turnos = 0
    
    for user, total_requests in users:
        # Count approved turnos
        approved_result = await db.execute(
            select(func.count(RequestItem.id))
            .join(Request)
            .where(
                Request.user_id == user.id,
                Request.created_at >= datetime.combine(start_date, datetime.min.time()),
                Request.created_at <= datetime.combine(end_date, datetime.max.time()),
                RequestItem.status == "approved"
            )
        )
        approved_turnos = approved_result.scalar() or 0
        
        # Count liberated turnos (status="cancelled" en BD)
        cancelled_result = await db.execute(
            select(func.count(RequestItem.id))
            .join(Request)
            .where(
                Request.user_id == user.id,
                Request.created_at >= datetime.combine(start_date, datetime.min.time()),
                Request.created_at <= datetime.combine(end_date, datetime.max.time()),
                RequestItem.status == "cancelled"
            )
        )
        cancelled_turnos = cancelled_result.scalar() or 0
        
        # Get last request date
        last_request_result = await db.execute(
            select(func.max(Request.created_at))
            .where(
                Request.user_id == user.id,
                Request.created_at >= datetime.combine(start_date, datetime.min.time()),
                Request.created_at <= datetime.combine(end_date, datetime.max.time())
            )
        )
        last_request = last_request_result.scalar()
        
        total_turnos += approved_turnos
        
        data.append([
            user.full_name,
            user.phone,
            "Activo" if user.is_active else "Inactivo",
            total_requests,
            approved_turnos,
            cancelled_turnos,
            last_request.strftime("%Y-%m-%d %H:%M") if last_request else "N/A"
        ])
    
    summary_data = {
        "Periodo": f"{start_date} a {end_date}",
        "Total Usuarios Activos": total_users,
        "Total Turnos Aprobados": total_turnos,
        "Promedio Turnos por Usuario": f"{(total_turnos / total_users):.1f}" if total_users > 0 else "0",
        "Fecha de Generación": datetime.now().strftime("%Y-%m-%d %H:%M")
    }
    
    headers = [
        "Nombre Completo",
        "Teléfono",
        "Estado",
        "Total Solicitudes",
        "Turnos Aprobados",
        "Turnos Liberados",
        "Última Solicitud"
    ]
    excel_file = create_excel_report(data, headers, "Reporte de Usuarios", summary_data)
    
    return Response(
        content=excel_file.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=usuarios_{start_date}_{end_date}.xlsx"}
    )

@router.get("/schedules")
async def get_schedules_report(
    start_date: date = Query(...),
    end_date: date = Query(...),
    admin: Admin = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db)
):
    """Export detailed schedules and capacity report"""
    result = await db.execute(
        select(Slot, Exhibitor, Schedule)
        .join(Exhibitor, Slot.exhibitor_id == Exhibitor.id)
        .join(Schedule, Slot.schedule_id == Schedule.id)
        .where(
            Slot.slot_date >= start_date,
            Slot.slot_date <= end_date
        )
        .order_by(Slot.slot_date, Exhibitor.name, Slot.start_ts)
    )
    slots = result.all()
    
    data = []
    total_slots = 0
    occupied_slots = 0
    
    for slot, exhibitor, schedule in slots:
        total_slots += 1
        
        # Count approved requests for this slot
        approved_result = await db.execute(
            select(func.count(RequestItem.id))
            .where(
                RequestItem.slot_id == slot.id,
                RequestItem.status == "approved"
            )
        )
        approved_count = approved_result.scalar() or 0
        
        capacity = slot.capacity or exhibitor.max_persons_per_slot or 5
        available = capacity - approved_count
        occupancy_pct = (approved_count / capacity * 100) if capacity > 0 else 0
        
        if approved_count > 0:
            occupied_slots += 1
        
        # Get weekday name
        weekday_names = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
        weekday = weekday_names[slot.slot_date.weekday()]
        
        status = "Completo" if approved_count >= capacity else "Disponible" if approved_count == 0 else "Parcial"
        
        data.append([
            exhibitor.code or "N/A",
            exhibitor.name,
            slot.slot_date.strftime("%Y-%m-%d"),
            weekday,
            slot.start_ts.strftime("%H:%M"),
            slot.end_ts.strftime("%H:%M"),
            capacity,
            approved_count,
            available,
            f"{occupancy_pct:.1f}%",
            status
        ])
    
    summary_data = {
        "Periodo": f"{start_date} a {end_date}",
        "Total Slots": total_slots,
        "Slots con Turnos": occupied_slots,
        "Slots Disponibles": total_slots - occupied_slots,
        "% Utilización": f"{(occupied_slots / total_slots * 100):.1f}%" if total_slots > 0 else "0%",
        "Fecha de Generación": datetime.now().strftime("%Y-%m-%d %H:%M")
    }
    
    headers = [
        "Código",
        "Exhibidor",
        "Fecha",
        "Día Semana",
        "Hora Inicio",
        "Hora Fin",
        "Capacidad",
        "Ocupados",
        "Disponibles",
        "% Ocupación",
        "Estado"
    ]
    excel_file = create_excel_report(data, headers, "Reporte de Horarios y Capacidad", summary_data)
    
    return Response(
        content=excel_file.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=horarios_capacidad_{start_date}_{end_date}.xlsx"}
    )

@router.get("/capacity")
async def get_capacity_report(
    start_date: date = Query(...),
    end_date: date = Query(...),
    admin: Admin = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db)
):
    """Export capacity and occupancy analysis report"""
    exhibitors_result = await db.execute(
        select(Exhibitor).where(Exhibitor.is_active == True).order_by(Exhibitor.name)
    )
    exhibitors = exhibitors_result.scalars().all()
    
    data = []
    total_capacity_all = 0
    total_occupied_all = 0
    
    for exhibitor in exhibitors:
        # Get all slots for this exhibitor in date range
        slots_result = await db.execute(
            select(Slot).where(
                Slot.exhibitor_id == exhibitor.id,
                Slot.slot_date >= start_date,
                Slot.slot_date <= end_date
            )
        )
        slots = slots_result.scalars().all()
        
        total_capacity = sum(slot.capacity or exhibitor.max_persons_per_slot or 5 for slot in slots)
        
        # Count total approved requests
        approved_result = await db.execute(
            select(func.count(RequestItem.id))
            .join(Slot)
            .where(
                Slot.exhibitor_id == exhibitor.id,
                Slot.slot_date >= start_date,
                Slot.slot_date <= end_date,
                RequestItem.status == "approved"
            )
        )
        total_occupied = approved_result.scalar() or 0
        
        total_available = total_capacity - total_occupied
        occupancy_rate = (total_occupied / total_capacity * 100) if total_capacity > 0 else 0
        
        total_capacity_all += total_capacity
        total_occupied_all += total_occupied
        
        # Determine status
        if occupancy_rate >= 90:
            status = "🔴 Completo"
        elif occupancy_rate >= 70:
            status = "🟠 Alto"
        elif occupancy_rate >= 40:
            status = "🟡 Medio"
        else:
            status = "🟢 Bajo"
        
        data.append([
            exhibitor.code or "N/A",
            exhibitor.name,
            len(slots),
            total_capacity,
            total_occupied,
            total_available,
            f"{occupancy_rate:.1f}%",
            status
        ])
    
    summary_data = {
        "Periodo": f"{start_date} a {end_date}",
        "Capacidad Total Sistema": total_capacity_all,
        "Espacios Ocupados": total_occupied_all,
        "Espacios Disponibles": total_capacity_all - total_occupied_all,
        "% Ocupación Global": f"{(total_occupied_all / total_capacity_all * 100):.1f}%" if total_capacity_all > 0 else "0%",
        "Fecha de Generación": datetime.now().strftime("%Y-%m-%d %H:%M")
    }
    
    headers = [
        "Código",
        "Exhibidor",
        "Total Slots",
        "Capacidad Total",
        "Espacios Ocupados",
        "Espacios Disponibles",
        "% Ocupación",
        "Estado"
    ]
    excel_file = create_excel_report(data, headers, "Análisis de Capacidad y Ocupación", summary_data)
    
    return Response(
        content=excel_file.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=capacidad_ocupacion_{start_date}_{end_date}.xlsx"}
    )

@router.get("/cancelled-by-user")
async def get_cancelled_by_user_report(
    start_date: date = Query(None),
    end_date: date = Query(None),
    admin: Admin = Depends(get_current_admin),
    db: AsyncSession = Depends(get_db)
):
    """Export report of liberated slots grouped by user"""
    # Si no se especifican fechas, usar todos los registros
    query = (
        select(
            RequestItem,
            Request,
            User,
            Slot,
            Exhibitor,
            Schedule
        )
        .join(Request, RequestItem.request_id == Request.id)
        .join(User, Request.user_id == User.id)
        .join(Slot, RequestItem.slot_id == Slot.id)
        .join(Exhibitor, Slot.exhibitor_id == Exhibitor.id)
        .outerjoin(Schedule, Slot.schedule_id == Schedule.id)
        .where(RequestItem.status == "cancelled")
    )
    
    # Filtrar por rango de fechas si se proporciona
    # Usar Request.created_at para coincidir con el dashboard (fecha de creación de la solicitud)
    # Si no se proporcionan fechas, mostrar todos los turnos liberados
    if start_date:
        query = query.where(Request.created_at >= datetime.combine(start_date, datetime.min.time()))
    if end_date:
        query = query.where(Request.created_at <= datetime.combine(end_date, datetime.max.time()))
    
    query = query.order_by(User.full_name, Slot.slot_date)
    
    result = await db.execute(query)
    rows = result.all()
    
    # Debug: verificar si hay turnos liberados sin filtro de fechas
    total_cancelled_all_time = 0
    if not rows and (start_date or end_date):
        # Intentar sin filtro de fechas para verificar si hay liberaciones
        debug_query = (
            select(func.count(RequestItem.id))
            .where(RequestItem.status == "cancelled")
        )
        debug_result = await db.execute(debug_query)
        total_cancelled_all_time = debug_result.scalar() or 0
    
    # Agrupar por usuario
    user_cancellations = {}
    total_cancellations = 0
    
    for row in rows:
        request_item, request, user, slot, exhibitor, schedule = row
        
        user_id = str(user.id)
        if user_id not in user_cancellations:
            user_cancellations[user_id] = {
                'user_name': user.full_name or 'Sin nombre',
                'user_phone': user.phone or 'Sin teléfono',
                'count': 0,
                'details': []
            }
        
        # Formatear horario
        if schedule:
            start_time = schedule.start_time.strftime("%H:%M") if schedule.start_time else "N/A"
            end_time = schedule.end_time.strftime("%H:%M") if schedule.end_time else "N/A"
            time_range = f"{start_time} - {end_time}"
        else:
            # Si no hay schedule, usar los timestamps del slot
            if slot.start_ts and slot.end_ts:
                start_time = slot.start_ts.strftime("%H:%M") if hasattr(slot.start_ts, 'strftime') else str(slot.start_ts)
                end_time = slot.end_ts.strftime("%H:%M") if hasattr(slot.end_ts, 'strftime') else str(slot.end_ts)
                time_range = f"{start_time} - {end_time}"
            else:
                time_range = "N/A"
        
        # Fecha de liberación (usar created_at del RequestItem cuando se cambió a cancelled)
        liberated_at = request_item.created_at
        liberated_date = liberated_at.strftime("%Y-%m-%d %H:%M") if liberated_at else "N/A"
        
        user_cancellations[user_id]['count'] += 1
        user_cancellations[user_id]['details'].append({
            'slot_date': slot.slot_date.strftime("%Y-%m-%d") if slot.slot_date else "N/A",
            'exhibitor_name': exhibitor.name or "N/A",
            'exhibitor_code': exhibitor.code or "N/A",
            'time_range': time_range,
            'liberated_date': liberated_date
        })
        total_cancellations += 1
    
    # Preparar datos para Excel
    data = []
    for user_id, user_data in sorted(user_cancellations.items(), key=lambda x: x[1]['count'], reverse=True):
        # Primera fila con resumen del usuario
        data.append([
            user_data['user_name'],
            user_data['user_phone'],
            user_data['count'],
            "",  # Fecha del turno (vacío en resumen)
            "",  # Exhibidor (vacío en resumen)
            "",  # Horario (vacío en resumen)
            ""   # Fecha de liberación (vacío en resumen)
        ])
        
        # Filas con detalles de cada liberación
        for detail in user_data['details']:
            data.append([
                "",  # Nombre (vacío para agrupar visualmente)
                "",  # Teléfono (vacío)
                "",  # Cantidad (vacío)
                detail['slot_date'],
                f"[{detail['exhibitor_code']}] {detail['exhibitor_name']}",
                detail['time_range'],
                detail['liberated_date']
            ])
    
    # Si no hay datos, agregar una fila indicando que no hay liberaciones
    if not data:
        periodo_texto = f"del {start_date} al {end_date}" if start_date and end_date else "en el sistema"
        data.append([
            f"No hay turnos liberados {periodo_texto}",
            "",
            "0",
            "",
            "",
            "",
            ""
        ])
    
    # Preparar resumen
    periodo_resumen = f"{start_date} a {end_date}" if start_date and end_date else "Todos los registros (sin filtro de fechas)"
    promedio = f"{(total_cancellations / len(user_cancellations)):.2f}" if user_cancellations and len(user_cancellations) > 0 else "0"
    
    summary_data = {
        "Periodo": periodo_resumen,
        "Total de Usuarios con Liberaciones": len(user_cancellations),
        "Total de Turnos Liberados": total_cancellations,
        "Promedio de Liberaciones por Usuario": promedio,
        "Fecha de Generación": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "Nota": "Si se especificaron fechas, se filtran por fecha de creación de la solicitud (Request.created_at)"
    }
    
    # Agregar información de debug si no hay resultados pero hay liberaciones en total
    if total_cancellations == 0 and total_cancelled_all_time > 0:
        summary_data["⚠️ Nota Importante"] = f"No se encontraron liberaciones en el rango de fechas, pero hay {total_cancelled_all_time} liberaciones en total. Intente sin filtro de fechas."
    
    headers = [
        "Nombre del Usuario",
        "Teléfono",
        "Cantidad de Liberaciones",
        "Fecha del Turno",
        "Exhibidor",
        "Horario",
        "Fecha de Liberación"
    ]
    
    excel_file = create_excel_report(data, headers, "Reporte de Turnos Liberados por Usuario", summary_data)
    
    filename = f"turnos_liberados_por_usuario_{start_date}_{end_date}.xlsx" if start_date and end_date else "turnos_liberados_por_usuario.xlsx"
    
    return Response(
        content=excel_file.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
